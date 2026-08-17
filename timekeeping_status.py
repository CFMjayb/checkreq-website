"""
timekeeping_status.py — Diocese-side Timekeeping status board, backfill
entry, and Excel export (added 2026-08-16, per Jay's direct request: "The
diocese mode will need screens to check the status of submissions for a
period, and the ability to enter missing data, and to generate a
spreadsheet with the data submitted for a period.").

New, single-purpose file per this project's standing modular-file
convention -- sibling to timekeeping.py / timekeeping_roster.py /
timekeeping_entries.py / timekeeping_review.py / timekeeping_roster_changes.py,
none of which needed to grow to support this. Delegates all grid-building
and hours-writing to timekeeping_entries.py's own
build_grid_context()/apply_hours_from_form() (extracted from that file this
same session specifically so this module wouldn't need a second copy) --
this file is ONLY the status board query, the diocese-side backfill
routes' own authorization/locking, and the Excel export.

THREE SCREENS, ALL GATED LIKE timekeeping_review.py's own _require_hr_admin
(hr_admin or beacon_admin at the diocese; a Cornerstone-Mode-selected
entity is rejected outright, same as every other diocese-wide Timekeeping
screen -- see that function's own docstring for the full reasoning, mirrored
here rather than imported, matching this codebase's own small-documented-
duplication precedent for permission-check helpers):

  1. GET  /admin/timekeeping/status/{period_id}
     The status board -- EVERY active parish under this diocese (via
     registry.list_parishes(), not just ones with a time_entry_submissions
     row), each with: submission status ("Not Started" when no row exists
     at all), total hours logged so far, and who/when submitted if
     applicable. Reachable from Payroll Periods' new "View Status" link per
     period row.

  2. GET/POST /admin/timekeeping/status/{period_id}/{parish_id}[/save]
     "Enter missing data" -- the diocese-side backfill grid for ONE
     specific parish/period, reusing timekeeping_entries.build_grid_context()/
     apply_hours_from_form() with diocese-side authorization instead of the
     parish's own can_manage_timekeeping() check. LOCK RULE (tightened
     2026-08-17 to Jay's three-state model): any period that is not 'open'
     refuses, diocese-side included and never bypassed; an already-'submitted'
     parish submission does NOT block the diocese -- "enter missing data"
     implies overriding a stuck
     state, so bypass_submission_lock=True is passed through. Every write
     is stamped edit_source='diocese' in portal.time_entry_edits (the CHECK
     constraint already allows this value, migration 038 -- no schema
     change needed), distinct from the parish's own 'parish'-sourced edits,
     so the audit trail always shows who really entered a given hour.

  3. GET  /admin/timekeeping/status/{period_id}/export
     Real portal.time_entries data for the whole period, across every
     parish in the diocese, as a downloadable .xlsx (openpyxl) with two
     sheets matching this org's own established Summary+Detail convention:
     Summary (one row per parish + staff + category, hours totaled for the
     whole period) and Detail (one row per day, for verification). See this
     module's own EXPORT COLUMN LAYOUT note below -- this is a first-cut
     guess, not yet confirmed against DME's real payroll-import process.

EXPORT COLUMN LAYOUT -- FLAGGED, NOT CONFIRMED (same open question already
on record in this project's CLAUDE.md/timekeeping build-plan history: "the
real payroll-export column layout is still unconfirmed with diocesan
staff"). The two-sheet Summary+Detail shape and the specific columns below
are this session's own best-effort match to how this codebase's other
Summary+Detail reports are built (Fund Summary Report's MTD/YTD tabs, QB
Time's Detail+Summary tabs) -- NOT verified against what DME's actual
payroll processor/software expects to import. Confirm with Jay/diocesan
staff before treating this as the final shape.
"""
from __future__ import annotations

import io

from fastapi import APIRouter, Request, Response
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db
import rbac
import registry
import cornerstone_mode
import timekeeping
import timekeeping_entries

router = APIRouter()

_current_user = None
_current_org = None
_render = None


def register(app, *, current_user, current_org, render) -> None:
    global _current_user, _current_org, _render
    _current_user, _current_org, _render = current_user, current_org, render
    app.include_router(router)


def _require_hr_admin(request: Request):
    """(user, org, None) when allowed, (None, None, response) when not.
    Mirrors timekeeping_review.py's own _require_hr_admin() (which itself
    mirrors timekeeping.py's _require_diocese_admin()) almost exactly --
    same "reject a Cornerstone-Mode-selected entity outright" rule, since
    every screen in this file reviews/edits diocese-wide payroll data, never
    a served parish's own linked org."""
    user = _current_user(request)
    if not user:
        return None, None, RedirectResponse("/login")
    org = _current_org(request)
    org_id = org["id"] if org else None
    if org_id is None or cornerstone_mode.is_cornerstone_org(org_id):
        return None, None, RedirectResponse("/portal")
    if not rbac.user_has_any_role(user["id"], ["hr_admin", "beacon_admin"], org_id=org_id):
        return None, None, JSONResponse({"error": "HR Admin access required"}, status_code=403)
    return user, org, None


def _get_period(period_id: int, org_id: int) -> dict | None:
    """Scoped lookup -- same choke-point discipline as registry.get_parish():
    returns None if period_id doesn't exist OR belongs to a different
    diocese, so a crafted id from another diocese's periods can never be
    viewed/edited/exported through this diocese's own routes."""
    return db.query_one(
        "SELECT * FROM portal.payroll_periods WHERE id = %s AND org_id = %s",
        (period_id, org_id),
    )


# ── Status board ─────────────────────────────────────────────────────────────

def list_parish_status_for_period(org_id: int, period_id: int) -> list[dict]:
    """Every ACTIVE parish under this diocese (registry.list_parishes(),
    not just ones with a submission row) with its submission status for
    this one period ("Not Started" is represented as submission_id IS
    NULL, rendered by the template, not a magic string here) and total
    hours logged so far. LEFT JOINs throughout so a parish with zero
    activity still gets a real row -- that's the whole point of this
    screen (surfacing who HASN'T started, not just who has)."""
    return db.query(
        """
        SELECT p.id AS parish_id, p.name AS parish_name, p.code AS parish_code,
               tes.id AS submission_id, tes.status AS submission_status,
               tes.submitted_by_user_id, tes.submitted_at,
               u.display_name AS submitted_by_name, u.email AS submitted_by_email,
               COALESCE(th.total_hours, 0) AS total_hours
          FROM portal.parishes p
          LEFT JOIN portal.time_entry_submissions tes
                 ON tes.parish_id = p.id AND tes.period_id = %(period_id)s
          LEFT JOIN checkreq.app_users u ON u.id = tes.submitted_by_user_id
          LEFT JOIN (
                SELECT sr.parish_id, SUM(te.hours) AS total_hours
                  FROM portal.time_entries te
                  JOIN portal.staff_roster sr ON sr.id = te.staff_id
                 WHERE te.period_id = %(period_id)s
                 GROUP BY sr.parish_id
          ) th ON th.parish_id = p.id
         WHERE p.org_id = %(org_id)s AND p.is_active
         ORDER BY p.name
        """,
        {"period_id": period_id, "org_id": org_id},
    )


@router.get("/admin/timekeeping/status", response_class=HTMLResponse)
def status_board_current(request: Request):
    """Top-level entry point (Administrative Tasks -> Timekeeping Status
    tile) -- jumps straight to the diocese's own current OPEN period's
    status board, the one parishes are actively working in right now, so
    an hr_admin doesn't have to go through Payroll Periods first. Falls
    back to Payroll Periods with a clear message if none is currently
    open. Distinct path-segment count from /status/{period_id} below, so
    there's no route-ordering ambiguity between the two."""
    user, org, err = _require_hr_admin(request)
    if err:
        return err
    period = timekeeping.get_current_open_period(org["id"])
    if not period:
        return RedirectResponse(
            "/admin/timekeeping/periods?error=No+open+payroll+period+right+now+%E2%80%94+open+one+first.",
            status_code=303,
        )
    return RedirectResponse(f"/admin/timekeeping/status/{period['id']}", status_code=303)


@router.get("/admin/timekeeping/status/{period_id}", response_class=HTMLResponse)
def status_board(period_id: int, request: Request):
    user, org, err = _require_hr_admin(request)
    if err:
        return err
    period = _get_period(period_id, org["id"])
    if not period:
        return RedirectResponse(
            "/admin/timekeeping/periods?error=That+payroll+period+was+not+found+for+this+diocese.",
            status_code=303,
        )
    rows = list_parish_status_for_period(org["id"], period_id)
    return _render(request, "timekeeping_status.html", user, {
        "current_org": org, "period": period, "rows": rows,
    })


# ── Excel export ─────────────────────────────────────────────────────────────
# NOTE ON PLACEMENT: this route MUST be registered before the
# /status/{period_id}/{parish_id} route below -- FastAPI/Starlette matches
# routes in registration order, and its path-matching happens BEFORE the
# {parish_id}: int type conversion is checked. Registering this route after
# the generic {parish_id} one would mean a GET to .../export always matches
# the {parish_id} pattern first and 422s trying to parse "export" as an
# int, never reaching this handler at all -- caught live via this session's
# own TestClient verification, not just reasoned about, before this
# placement was settled on.

_SUMMARY_SQL = """
    SELECT p.code AS parish_code, p.name AS parish_name,
           sr.last_name, sr.first_name, sr.position, sr.employee_number,
           c.label AS category_label,
           SUM(te.hours) AS total_hours
      FROM portal.time_entries te
      JOIN portal.staff_roster sr ON sr.id = te.staff_id
      JOIN portal.parishes p ON p.id = sr.parish_id
      JOIN portal.timekeeping_categories c ON c.id = te.category_id
     WHERE te.period_id = %(period_id)s AND p.org_id = %(org_id)s
     GROUP BY p.code, p.name, sr.last_name, sr.first_name, sr.position, sr.employee_number,
              c.label, c.sort_order
    HAVING SUM(te.hours) > 0
     ORDER BY p.name, sr.last_name, sr.first_name, c.sort_order
"""

_DETAIL_SQL = """
    SELECT p.code AS parish_code, p.name AS parish_name,
           sr.last_name, sr.first_name,
           c.label AS category_label, te.work_date, te.hours
      FROM portal.time_entries te
      JOIN portal.staff_roster sr ON sr.id = te.staff_id
      JOIN portal.parishes p ON p.id = sr.parish_id
      JOIN portal.timekeeping_categories c ON c.id = te.category_id
     WHERE te.period_id = %(period_id)s AND p.org_id = %(org_id)s AND te.hours > 0
     ORDER BY p.name, sr.last_name, sr.first_name, te.work_date, c.sort_order
"""

_HEADER_FILL = "1F3D2E"  # --color-franciscan-green, this app's own brand token


def _write_header(ws, headers: list[str]) -> None:
    fill = PatternFill(start_color=_HEADER_FILL, end_color=_HEADER_FILL, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_export_workbook(org: dict, period: dict) -> bytes:
    """Real portal.time_entries data for one period, across every parish in
    this diocese -- see the module docstring's EXPORT COLUMN LAYOUT note.
    Only rows with real (>0) hours are included in either sheet -- a sparse
    "every staff x every category" grid would mostly be empty and adds
    nothing a reviewer needs."""
    summary_rows = db.query(_SUMMARY_SQL, {"period_id": period["id"], "org_id": org["id"]})
    detail_rows = db.query(_DETAIL_SQL, {"period_id": period["id"], "org_id": org["id"]})

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, [
        "Parish Code", "Parish Name", "Last Name", "First Name", "Position",
        "Employee #", "Category", "Total Hours",
    ])
    for r_i, row in enumerate(summary_rows, start=2):
        ws.cell(row=r_i, column=1, value=row["parish_code"])
        ws.cell(row=r_i, column=2, value=row["parish_name"])
        ws.cell(row=r_i, column=3, value=row["last_name"])
        ws.cell(row=r_i, column=4, value=row["first_name"])
        ws.cell(row=r_i, column=5, value=row["position"])
        ws.cell(row=r_i, column=6, value=row["employee_number"])
        ws.cell(row=r_i, column=7, value=row["category_label"])
        ws.cell(row=r_i, column=8, value=float(row["total_hours"])).number_format = "0.00"
    _set_widths(ws, [12, 32, 16, 14, 20, 14, 14, 12])
    if not summary_rows:
        ws.cell(row=2, column=1, value="No hours logged for this period yet.")

    ws2 = wb.create_sheet("Detail")
    _write_header(ws2, [
        "Parish Code", "Parish Name", "Last Name", "First Name",
        "Category", "Date", "Hours",
    ])
    for r_i, row in enumerate(detail_rows, start=2):
        ws2.cell(row=r_i, column=1, value=row["parish_code"])
        ws2.cell(row=r_i, column=2, value=row["parish_name"])
        ws2.cell(row=r_i, column=3, value=row["last_name"])
        ws2.cell(row=r_i, column=4, value=row["first_name"])
        ws2.cell(row=r_i, column=5, value=row["category_label"])
        d = row["work_date"]
        ws2.cell(row=r_i, column=6, value=d.isoformat() if hasattr(d, "isoformat") else str(d))
        ws2.cell(row=r_i, column=7, value=float(row["hours"])).number_format = "0.00"
    _set_widths(ws2, [12, 32, 16, 14, 14, 12, 10])
    if not detail_rows:
        ws2.cell(row=2, column=1, value="No hours logged for this period yet.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/admin/timekeeping/status/{period_id}/export")
def export_period(period_id: int, request: Request):
    user, org, err = _require_hr_admin(request)
    if err:
        return err
    period = _get_period(period_id, org["id"])
    if not period:
        return JSONResponse({"error": "That payroll period was not found for this diocese."}, status_code=404)

    content = build_export_workbook(org, period)
    label = (period.get("label") or f"{period['period_start']}_to_{period['period_end']}")
    safe_label = "".join(c if c.isalnum() or c in " -_." else "_" for c in str(label)).strip() or "Period"
    filename = f"{org['code']} Timekeeping {safe_label}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Diocese-side backfill entry ─────────────────────────────────────────────

@router.get("/admin/timekeeping/status/{period_id}/{parish_id}", response_class=HTMLResponse)
def backfill_entry_page(period_id: int, parish_id: int, request: Request):
    user, org, err = _require_hr_admin(request)
    if err:
        return err
    period = _get_period(period_id, org["id"])
    if not period:
        return RedirectResponse(
            "/admin/timekeeping/periods?error=That+payroll+period+was+not+found+for+this+diocese.",
            status_code=303,
        )
    parish = registry.get_parish(parish_id, org["id"])
    if not parish:
        return RedirectResponse(
            f"/admin/timekeeping/status/{period_id}?error=That+parish+was+not+found+for+this+diocese.",
            status_code=303,
        )
    # can_enter=True: the route gate above (hr_admin/beacon_admin at this
    # diocese) IS the authorization for this screen -- unlike the parish's
    # own grid, this isn't re-derived from can_manage_timekeeping() (which
    # doesn't even check the hr_admin role, since that role didn't exist
    # when can_manage_timekeeping() was written for the parish/roster
    # side). bypass_submission_lock=True: an already-'submitted' parish
    # submission never blocks the diocese from entering missing data
    # (Jay's own explicit ask) -- only a genuinely 'processed' period does,
    # which build_grid_context() enforces regardless of this flag.
    ctx = timekeeping_entries.build_grid_context(
        parish, org["id"], period, can_enter=True, bypass_submission_lock=True,
    )
    ctx["diocese_mode"] = True
    ctx["back_link_url"] = f"/admin/timekeeping/status/{period_id}"
    ctx["back_link_label"] = "Back to Status Board"
    ctx["save_url"] = f"/admin/timekeeping/status/{period_id}/{parish_id}/save"
    return _render(request, "timekeeping_entry.html", user, ctx)


@router.post("/admin/timekeeping/status/{period_id}/{parish_id}/save")
async def backfill_entry_save(period_id: int, parish_id: int, request: Request):
    user, org, err = _require_hr_admin(request)
    if err:
        return err
    period = _get_period(period_id, org["id"])
    if not period:
        return RedirectResponse(
            "/admin/timekeeping/periods?error=That+payroll+period+was+not+found+for+this+diocese.",
            status_code=303,
        )
    parish = registry.get_parish(parish_id, org["id"])
    if not parish:
        return RedirectResponse(
            f"/admin/timekeeping/status/{period_id}?error=That+parish+was+not+found+for+this+diocese.",
            status_code=303,
        )
    if period["status"] != "open":
        # The one lock this route can never override -- only an 'open' period
        # is writable, diocese-side included (Jay, 2026-08-17: "only 'open' is
        # available for data entry and submission"). Tightened from the old
        # `== "processed"` test, which left 'closed'/'future' periods writable
        # through this backfill route. NOTE this is a deliberate narrowing of
        # what the diocese can do: an hr_admin can no longer backfill hours
        # into a closed period at all. If that turns out to be needed for a
        # genuine late correction, this is the one line to revisit -- it is
        # called out in this session's write-up rather than left as a surprise.
        return RedirectResponse(
            f"/admin/timekeeping/status/{period_id}/{parish_id}?error=notopen",
            status_code=303,
        )
    # Make sure a submission row exists before writing hours against it --
    # same reason the parish's own /timekeeping/save does this (so the
    # status board always has something to show the moment ANY hours are
    # logged, diocese-entered or not).
    timekeeping_entries.get_or_create_submission(period["id"], parish["id"])
    form = await request.form()
    saved = timekeeping_entries.apply_hours_from_form(
        period, parish, org["id"], user, form, edit_source="diocese",
    )
    return RedirectResponse(
        f"/admin/timekeeping/status/{period_id}/{parish_id}?saved={saved}", status_code=303
    )
